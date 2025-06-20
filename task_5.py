from openpyxl import load_workbook
import math

wb = load_workbook("sagatave_eksamenam.xlsx", data_only=True)
ws = wb["Lapa_0"]
max_row = ws.max_row

count_total_customer_amount = 0
for row in range(2, max_row + 1):
    total = ws[f"N{row}"].value
    customer = ws[f"F{row}"].value
    amount = ws[f"L{row}"].value

    # Task 5
    if customer == "Korporatīvais" and amount and 40 <= amount <= 50:
        count_total_customer_amount += total

count_total_customer_amount_floored = math.floor(count_total_customer_amount)

print(count_total_customer_amount_floored)
