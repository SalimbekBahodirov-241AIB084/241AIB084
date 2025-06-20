from openpyxl import load_workbook
import math

wb = load_workbook("sagatave_eksamenam.xlsx", data_only=True)
ws = wb["Lapa_0"]
max_row = ws.max_row

count_product_price = []
for row in range(2, max_row + 1):
    product = ws[f"I{row}"].value
    price = ws[f"K{row}"].value

    # Task 4
    if product and "LaserJet" in product:
        count_product_price.append(price)

avg_product_price = math.floor(sum(count_product_price) / len(count_product_price))

print(avg_product_price)
