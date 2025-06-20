from openpyxl import load_workbook


wb = load_workbook("sagatave_eksamenam.xlsx", data_only=True)
ws = wb["Lapa_0"]
max_row = ws.max_row

count_address_city = 0
for row in range(2, max_row + 1):
    address = ws[f"D{row}"].value
    city = ws[f"E{row}"].value

    # Task 3
    if address == "Adulienas iela" and city in ("Valmiera", "Saulkrasti"):
        count_address_city += 1

print(count_address_city)
