from openpyxl import load_workbook


wb = load_workbook("sagatave_eksamenam.xlsx", data_only=True)
ws = wb["Lapa_0"]
max_row = ws.max_row

count_priority_date = 0
for row in range(2, max_row + 1):
    priority = ws[f"H{row}"].value
    delivery_date = ws[f"J{row}"].value

    # Task 2
    if priority == "High" and delivery_date.year == 2015:
        count_priority_date += 1

print(count_priority_date)
