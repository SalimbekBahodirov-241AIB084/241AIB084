from openpyxl import load_workbook


wb = load_workbook("sagatave_eksamenam.xlsx", data_only=True)
ws = wb["Lapa_0"]
max_row = ws.max_row

count_address_amount = 0
for row in range(2, max_row + 1):
    address = ws[f"D{row}"].value
    amount = ws[f"L{row}"].value

    # Task 1
    if address and address.startswith("Ain") and amount < 40:
        count_address_amount += 1

print(count_address_amount)
